from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
body_font = Font(name='Arial', size=10)
center = Alignment(horizontal='center', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='AAAAAA'),
    right=Side(style='thin', color='AAAAAA'),
    top=Side(style='thin', color='AAAAAA'),
    bottom=Side(style='thin', color='AAAAAA')
)
data_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

def header_row(ws, row, values, bg):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i+1, value=v)
        c.font = header_font
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = center
        c.border = thin_border

def data_row(ws, row, values):
    bg = 'F5F5F5' if row % 2 == 0 else 'FFFFFF'
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i+1, value=str(v) if v is not None else '')
        c.font = body_font
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = center
        c.border = data_border

def section_title(ws, row, cols, text, bg):
    ws.merge_cells(f'A{row}:{get_column_letter(cols)}{row}')
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = center
    ws.row_dimensions[row].height = 22

# ===== Sheet 1: 银行卡 =====
ws1 = wb.active
ws1.title = '银行卡'

ws1.merge_cells('A1:F1')
ws1['A1'].value = '投资模拟器 - 全部银行卡账户清单'
ws1['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
ws1['A1'].fill = PatternFill('solid', start_color='1A237E')
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 30

# 内地银行
section_title(ws1, 2, 6, '🇨🇳 内地银行 (人民币 CNY)', 'B71C1C')
headers = ['序号', '银行名称', '卡号', '持卡人', '备注', '地区']
header_row(ws1, 3, headers, '4A148C')
ws1.row_dimensions[3].height = 20

mainland_cards = [
    (1, '工商银行 ICBC', '6222356126761078686', '复旦大学', '储蓄卡', '内地'),
    (2, '工商银行 ICBC', '6257091287019652', '中国人民大学', '储蓄卡', '内地'),
    (3, '工商银行 ICBC', '6257608303684569', '北京师范大学', '储蓄卡', '内地'),
    (4, '工商银行 ICBC', '6256960555482116', '中国农业大学', '储蓄卡', '内地'),
    (5, '工商银行 ICBC', '6258936136932259', '东莞市第二高级中学', '储蓄卡', '内地'),
    (6, '工商银行 ICBC', '427721947812932033', '冯程轩', '储蓄卡', '内地'),
    (7, '建设银行 CCB', '6222802225595725722', '上海交通大学', '储蓄卡', '内地'),
    (8, '建设银行 CCB', '6993475370026539779', '北京航空航天大学', '储蓄卡', '内地'),
    (9, '建设银行 CCB', '6993477499972608074', '中国科学院大学', '储蓄卡', '内地'),
    (10, '建设银行 CCB', '6993478055491680481', '北京科技大学', '储蓄卡', '内地'),
    (11, '建设银行 CCB', '6993477200026359281', '东莞市第七高级中学', '储蓄卡', '内地'),
    (12, '中国银行 BOC', '6222809535503965036', '同济大学', '储蓄卡', '内地'),
    (13, '中国银行 BOC', '6857949556405107945', '北京理工大学', '储蓄卡', '内地'),
    (14, '中国银行 BOC', '6857945236057598066', '北京协和医学院', '储蓄卡', '内地'),
    (15, '中国银行 BOC', '6857941846317114624', '北京交通大学', '储蓄卡', '内地'),
    (16, '中国银行 BOC', '6857948273017407128', '东莞市第八高级中学', '中银账户①', '内地'),
    (17, '中国银行 BOC', '8727503248106057', '东莞市第八高级中学', '中银账户②', '内地'),
    (18, '农业银行 ABC', '6222808480706815121', '华东师范大学', '储蓄卡', '内地'),
    (19, '农业银行 ABC', '6222808480706815121', '上海财经大学', '储蓄卡', '内地'),
]
for i, row in enumerate(mainland_cards):
    data_row(ws1, 4+i, row)

# 香港银行
r = 4 + len(mainland_cards)
section_title(ws1, r, 6, '🇭🇰 香港银行 (港元 HKD / 美元 USD)', '006064')
header_row(ws1, r+1, headers, '00838F')

hk_cards = [
    (20, '汇丰银行 HSBC', '6222807846702561206', '复旦大学', 'HKD储蓄', '香港'),
    (21, '汇丰银行 HSBC', '6222807299293970759', '华东师范大学', 'HKD储蓄', '香港'),
    (22, '汇丰银行 HSBC', '245941620609428422', '东莞市第二高级中学', 'HKD信用卡', '香港'),
    (23, '恒生银行 Hang Seng', '6222808932201878052', '同济大学', 'HKD储蓄', '香港'),
    (24, '中银香港 BOCHK', '6222802706288717432', '上海交通大学', 'HKD储蓄', '香港'),
    (25, '渣打银行 SC', '69852936050580755', '东莞市第七高级中学', 'HKD储蓄', '香港'),
]
for i, row in enumerate(hk_cards):
    data_row(ws1, r+2+i, row)

# 英国银行
r2 = r + 2 + len(hk_cards)
section_title(ws1, r2, 6, '🇬🇧 英国银行 (美元 USD / 英镑 GBP)', '1B5E20')
header_row(ws1, r2+1, headers, '2E7D32')

uk_cards = [
    (26, '英格兰银行 BoE', '62193400809416219531', '叶珒铭', 'GBP Current', '英国'),
    (27, '英格兰银行 BoE', '62193401309416219531', '叶珒铭', 'GBP Savings', '英国'),
    (28, 'HSBC UK', '8793400209413219531', '叶珒铭', 'USD Account', '英国'),
    (29, 'HSBC UK', '8793403509413219531', '叶珒铭', 'USD Premier', '英国'),
]
for i, row in enumerate(uk_cards):
    data_row(ws1, r2+2+i, row)

for i, w in enumerate([8, 24, 26, 22, 16, 8]):
    ws1.column_dimensions[get_column_letter(i+1)].width = w

# ===== Sheet 2: 证券账户 =====
ws2 = wb.create_sheet('证券账户')

ws2.merge_cells('A1:D1')
ws2['A1'].value = '投资模拟器 - 全部证券账户清单'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
ws2['A1'].fill = PatternFill('solid', start_color='1A237E')
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 30

sec_headers = ['序号', '证券账户号', '持有人', '备注']
section_title(ws2, 2, 4, '🏛️ 中信证券 CITIC Securities', 'E65100')
header_row(ws2, 3, sec_headers, 'BF360C')
ws2.row_dimensions[3].height = 20

citics_accounts = [
    (1, 'A88810001', '叶珒铭', '主账户'),
    (2, 'A88820001', '东莞市第二高级中学', ''),
    (3, 'A88830001', '复旦大学', ''),
    (4, 'A88840001', '上海交通大学', ''),
]
for i, row in enumerate(citics_accounts):
    data_row(ws2, 4+i, row)

r3 = 4 + len(citics_accounts)
section_title(ws2, r3, 4, '💎 高盛证券 Goldman Sachs', '4A148C')
header_row(ws2, r3+1, sec_headers, '6A1B9A')

goldman_accounts = [
    (5, 'G90010001', '叶珒铭', '主账户'),
    (6, 'G90020001', '华东师范大学', ''),
]
for i, row in enumerate(goldman_accounts):
    data_row(ws2, r3+2+i, row)

for i, w in enumerate([8, 20, 22, 12]):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# ===== Sheet 3: 汇总 =====
ws3 = wb.create_sheet('汇总')

ws3.merge_cells('A1:C1')
ws3['A1'].value = '账户总览汇总'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
ws3['A1'].fill = PatternFill('solid', start_color='1A237E')
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 30

sum_rows = [
    ('类别', '数量', '说明'),
    ('内地银行卡', 19, 'CNY 人民币'),
    ('香港银行卡', 6, 'HKD/USD'),
    ('英国银行卡', 4, 'USD/GBP'),
    ('银行卡合计', '=B3+B4+B5', ''),
    ('中信证券账户', 4, 'A股/港股'),
    ('高盛证券账户', 2, '港股/美股'),
    ('证券账户合计', '=B7+B8', ''),
    ('全部账户合计', '=B6+B9', ''),
]

for i, row in enumerate(sum_rows):
    for j, v in enumerate(row):
        c = ws3.cell(row=i+2, column=j+1, value=v)
        if i == 0:
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='283593')
        elif i in [3, 6, 7]:
            c.font = Font(name='Arial', bold=True, size=10, color='1A237E')
            c.fill = PatternFill('solid', start_color='E8EAF6')
        else:
            c.font = body_font
            c.fill = PatternFill('solid', start_color='F5F5F5' if i % 2 == 0 else 'FFFFFF')
        c.alignment = center
        c.border = thin_border

for i, w in enumerate([18, 12, 16]):
    ws3.column_dimensions[get_column_letter(i+1)].width = w

wb.save('账户清单.xlsx')
print('Done')
